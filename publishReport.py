import os
import openpyxl
from openpyxl import Workbook
from printToText import printToFile
from printToText import createReadMe
from main import main
def create_report(results,template_name='NONE',sheet='NONE', map = 'NONE',scan_type=['Adult','Abdomen']):
    if template_name=='NONE':
        wb = Workbook()
        ws = wb.active
        ws.title = 'CT_IQ'
    else:
        wb = openpyxl.load_workbook(template_name)
        ws = wb[sheet]

    if map != 'NONE':
        map = getMapping(map)

    ct_number_accuracy(ws,results,scan_type,template_name,map[0])
    low_contrast_resolution(ws,results,scan_type,template_name,map[1])
    ct_number_uniformity(ws,results,scan_type,template_name,map[2])
    high_contrast_resolution(ws, results, scan_type,template_name,map[3])

    # checks if the directory qcReport exists
    if not os.path.exists("qcReport"):
        # if the folder is not present, then create it
        os.makedirs("qcReport")

    # print('Absolute path of file:     ',
    #       os.path.abspath(__file__))
    # print('Absolute directoryname: ',
    #       os.path.dirname(os.path.abspath(__file__)))

    createReadMe()

    if template_name == 'NONE':
        wb.save('NewReport.xlsx')
    else:
        wb.save(template_name)

def getMapping(filename):
    f = open(filename)
    temp = f.read().splitlines()
    m1 = temp[0].split(",")
    m2 = temp[1].split(",")
    m3 = temp[2].split(",")
    m4 = temp[3].split(",")

    return m1, m2, m3, m4

def ct_number_accuracy(ws,results,scan_type,template='NONE', map='NONE'):
    printToFile(results[0])

    if template == 'NONE' and map == 'NONE':
        ws.cell(row=1,column=1).value = 'CT Number Accuracy'
        ws.cell(row=2, column=1).value = 'Material'
        ws.cell(row=2, column=2).value = 'HU'
        ws.cell(row=2, column=3).value = 'Pass/Fail'
        ws.cell(row=3,column=1).value = 'Water'
        ws.cell(row=4, column=1).value = 'Air'
        ws.cell(row=5, column=1).value = 'Bone'
        ws.cell(row=6, column=1).value = 'Poly'
        ws.cell(row=7, column=1).value = 'Acrylic'

        row_start = 3
        row_pos = 3
        col_pos = 2

        insert_sel = 0
        for i in range (row_start,row_start+5):
            ws.cell(row=row_pos,column=col_pos).value = results[0][insert_sel][0]
            row_pos += 1
            insert_sel += 1

    if template != 'NONE' and map != 'NONE':
        j = 0
        for i in map:
            ws[i].value = results[0][j][0]
            j += 1
    else:
        if scan_type[0] == 'Adult':
            row_start = 17
            row_pos = 17
        if scan_type[0] == 'Pediatric':
            row_start = 24
            row_pos = 24

        if scan_type[1] == 'Abdomen':
            col_pos = 2
        if scan_type[1] == 'Head':
            col_pos = 8

        insert_sel = 0
        for i in range (row_start,row_start+5):
            ws.cell(row=row_pos,column=col_pos).value = results[0][insert_sel][0]
            row_pos += 1
            insert_sel += 1

def low_contrast_resolution(ws,results,scan_type,template='NONE', map='NONE'):
    printToFile(results[1])

    if template == 'NONE' and map == 'NONE':
        if scan_type[0] == 'Adult' and scan_type[1] == 'Abdomen':
            col_start = 3
            col_pos = 3
            row_pos = 44
        if scan_type[0] == 'Adult' and scan_type[1] == 'Head':
            col_start = 3
            col_pos = 3
            row_pos = 45
        if scan_type[0] == 'Pediatric' and scan_type[1] == 'Abdomen':
            col_start = 3
            col_pos = 3
            row_pos = 46
        if scan_type[0] == 'Pediatric' and scan_type[1] == 'Head':
            col_start = 3
            col_pos = 3
            row_pos = 47

        roi_val = 0
        for i in range(col_start, col_start + 3):
            ws.cell(row=row_pos, column=col_pos).value = results[1][2][roi_val]
            col_pos += 1
            roi_val += 1

    if template != 'NONE' and map != 'NONE':
        roi_val = 0
        for i in map:
            ws[i].value = results[1][2][roi_val]
            roi_val += 1

def ct_number_uniformity(ws,results,scan_type,template='NONE', map='NONE'):
    printToFile(results[2])
    if template == 'NONE' and map == 'NONE':
        if scan_type[0] == 'Adult' and scan_type[1] == 'Abdomen':
            col_start = 3
            col_pos = 3
            row_pos = 51
        if scan_type[0] == 'Adult' and scan_type[1] == 'Head':
            col_start = 3
            col_pos = 3
            row_pos = 55
        if scan_type[0] == 'Pediatric' and scan_type[1] == 'Abdomen':
            col_start = 3
            col_pos = 3
            row_pos = 60
        if scan_type[0] == 'Pediatric' and scan_type[1] == 'Head':
            col_start = 3
            col_pos = 3
            row_pos = 64

        roi_val = 0
        for i in range (col_start,col_start+5):
            ws.cell(row=row_pos,column=col_pos).value = results[2][1][roi_val]
            col_pos += 1
            roi_val += 1

    if template != 'NONE' and map != 'NONE':
        roi_val = 0
        for i in map:
            ws[i].value = results[2][1][roi_val]
            roi_val += 1

def high_contrast_resolution(ws,results,scan_type,template='NONE', map='NONE'):
    # printToFile(results[3])

    if template == 'NONE' and map == 'NONE':
        if scan_type[0] == 'Adult' and scan_type[1] == 'Abdomen':
            col_pos = 2
            row_pos = 71
        if scan_type[0] == 'Adult' and scan_type[1] == 'Head':
            col_pos = 7
            row_pos = 74
        if scan_type[0] == 'Pediatric' and scan_type[1] == 'Abdomen':
            col_pos = 2
            row_pos = 71
        if scan_type[0] == 'Pediatric' and scan_type[1] == 'Head':
            col_pos = 7
            row_pos = 74

        ws.cell(row=row_pos,column=col_pos).value = results[3][0]

    if template != 'NONE' and map != 'NONE':
        for i in map:
            ws[i].value = results[3][0]

if __name__ == "__main__":
    data_path = r"C:\Users\derek\Documents\RAMD5394IndependentStudy\ACRQCtool\ADULT_ABD_10\\"
    test_results = main(data_path)
    create_report(test_results,'template.xlsx','CT_IQ', 'reportMapping.txt',['Adult','Abdomen'])
    # create_report(test_results)